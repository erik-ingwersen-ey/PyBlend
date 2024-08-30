import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder


def create_formatted_excel(
    stockpiles_df: pd.DataFrame,
    engines_df: pd.DataFrame,
    operations_df: pd.DataFrame,
    outputs_df_out: pd.DataFrame,
    file_name: str = "optimization_output.xlsx",
):
    """Create a formatted Excel file from the provided DataFrames.

    This function takes multiple pandas DataFrames containing stockpile,
    engine, operation, and output information, and creates a well-formatted
    Excel file. Each DataFrame is added to a separate sheet within the Excel
    workbook. The function allows for customization of formatting, such as
    alternating row colors and bolding the last row of data.

    Args:
        stockpiles_df (pd.DataFrame): A `pandas.DataFrame` containing stockpile information.
        engines_df (pd.DataFrame): A `pandas.DataFrame` containing engine information.
        operations_df (pd.DataFrame): A `pandas.DataFrame` containing operation details.
        outputs_df_out (pd.DataFrame): A `pandas.DataFrame` containing output information from the
            optimization.
        file_name (str?): Name of the Excel file to be created. Defaults to
            'optimization_output.xlsx'.
    """
    # Create a new workbook and remove the default sheet
    wb = Workbook()
    wb.remove(wb.active)

    def add_dataframe_to_sheet(
        wb,
        df,
        sheet_name,
        color_alternating_rows=False,
        color_by_engine=False,
        bold_last_row=False,
    ):
        """Add a DataFrame to an Excel sheet with optional formatting.

        This helper function creates a new sheet in the provided workbook and
        populates it with the data from the given DataFrame. It supports various
        formatting options, including alternating row colors, coloring by engine
        type, and bolding the last row. The function also adjusts column widths
        based on the content of the cells and hides gridlines for a cleaner
        appearance.

        Args:
            wb (Workbook): The workbook to which the sheet is added.
            df (pd.DataFrame): The DataFrame to be added.
            sheet_name (str): The name of the sheet.
            color_alternating_rows (bool?): Whether to color rows with alternating colors. Defaults to False.
            color_by_engine (bool?): Whether to color rows by different engines. Defaults to False.
            bold_last_row (bool?): Whether to make the last row bold. Defaults to False.
        """
        ws = wb.create_sheet(title=sheet_name)
        colors = ["B9C8DE", "DEE6F0"]  # Rotation colors for alternating rows

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, list):
                    value = str(value)
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                # Header row
                if r_idx == 1:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.fill = PatternFill(
                        start_color="5B80B8", end_color="5B80B8", fill_type="solid"
                    )
                # Data rows
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(value, (int, float)) and abs(value) >= 1000:
                        cell.number_format = "#,##0"

                    # Apply row coloring
                    if color_alternating_rows:
                        # r_idx - 2 to account for the header row
                        color_index = (r_idx - 2) % len(colors)
                        cell.fill = PatternFill(
                            start_color=colors[color_index],
                            end_color=colors[color_index],
                            fill_type="solid",
                        )
                    elif color_by_engine and "engine" in df.columns:
                        engine_colors = {
                            1: "68B6F2",  # Blue 1
                            2: "5C86DA",  # Blue 2
                            3: "525DB6",  # Blue 3
                            4: "1E1D39",  # Blue 4
                            5: "1E1D39",  # Blue 5
                        }
                        eng_value = df.loc[r_idx - 2, "engine"]
                        if hasattr(eng_value, "values"):
                            eng_value = sum(eng_value.values)
                        if isinstance(eng_value, list):
                            eng_value = sum(eng_value)

                        fill_color = engine_colors.get(eng_value, "5988B8")
                        cell.fill = PatternFill(
                            start_color=fill_color,
                            end_color=fill_color,
                            fill_type="solid",
                        )
                        if fill_color in ["525DB6", "1E1D39"]:
                            cell.font = Font(color="FFFFFF")

            # Make the last row bold
            if bold_last_row and r_idx == len(df) + 1:
                for cell in ws[r_idx]:
                    if cell.fill.bgColor.value.strip("0") in ["525DB6", "1E1D39"]:
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.font = Font(bold=True)

        # Adjust column widths
        dim_holder = DimensionHolder(worksheet=ws)
        for col_idx, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) for cell in col)
            dim_holder[get_column_letter(col_idx)] = ColumnDimension(
                ws, min=col_idx, max=col_idx, width=max_length + 2
            )
        ws.column_dimensions = dim_holder

        # Hide gridlines
        ws.sheet_view.showGridLines = False

    # Add DataFrames to sheets
    add_dataframe_to_sheet(wb, stockpiles_df, "Stockpiles", color_alternating_rows=True)
    add_dataframe_to_sheet(wb, engines_df, "Engines")
    add_dataframe_to_sheet(
        wb, operations_df, "Operations", color_by_engine=True, bold_last_row=True
    )
    add_dataframe_to_sheet(wb, outputs_df_out, "Outputs")

    # Save the workbook
    wb.save(file_name)
    print(f"Excel file '{file_name}' created successfully.")


# Convert input and output JSON files to dataframes
def explode_quality_rows(
    df: pd.DataFrame,
    quality_col_prefix: str = "quality_",
) -> pd.DataFrame:
    """Explode the 'quality_*' columns into rows.

    This function takes a DataFrame containing columns that start with a
    specified prefix (default is 'quality_') and transforms those columns,
    which contain dictionaries of quality parameters, into separate rows.
    Each dictionary will be expanded into its own set of columns, while the
    other columns in the DataFrame will be repeated for each exploded row.
    This is useful for normalizing data structures where quality parameters
    are stored in a nested format.

    Args:
        df (pd.DataFrame): The dataframe with the output specifications containing
            the quality columns that hold dictionaries of quality parameters to be
            extracted.
        quality_col_prefix (str?): The prefix of the quality columns that
            contain the output specifications to be extracted into different
            columns.
            Defaults to 'quality_'.

    Returns:
        pd.DataFrame: A new DataFrame with the quality dictionaries extracted into
            new columns, where each dictionary is represented as a separate row.
    """
    # Create a new DataFrame to store exploded rows
    exploded_df = pd.DataFrame()

    for index, row in df.iterrows():
        quality_dict_list = [
            value for key, value in row.items() if key.startswith(quality_col_prefix)
        ]

        # Convert the list of quality dictionaries into a DataFrame
        quality_df = pd.DataFrame(quality_dict_list)

        # Repeat the original columns for each exploded row
        repeated_columns = pd.DataFrame(
            [
                row.drop(
                    labels=[
                        col for col in df.columns if col.startswith(quality_col_prefix)
                    ]
                )
            ]
            * len(quality_df)
        )

        # Concatenate the repeated columns with the quality columns
        exploded_row_df = pd.concat(
            [
                repeated_columns.reset_index(drop=True),
                quality_df.reset_index(drop=True),
            ],
            axis=1,
        )

        # Append to the exploded DataFrame
        exploded_df = pd.concat([exploded_df, exploded_row_df], ignore_index=True)

    return exploded_df


def assign_engines_to_stockpiles(
    stockpiles_df: pd.DataFrame, engines_df: pd.DataFrame
) -> pd.DataFrame:
    """Assign engines to stockpiles based on matching yards and rails.

    This function takes two pandas DataFrames: one containing stockpile
    information and the other containing engine information. It iterates
    through the stockpiles and assigns engines to each stockpile based on
    whether the stockpile's yard matches any of the engine's yards and if
    the engine's rail is present in the stockpile's rails. The result is an
    updated stockpile DataFrame that includes an 'engines' column, which
    lists the IDs of the assigned engines for each stockpile.

    Args:
        stockpiles_df (pd.DataFrame): A `pandas.DataFrame` containing stockpile
            information including 'rails' and 'yard'.
        engines_df (pd.DataFrame): A `pandas.DataFrame` containing engine
            information including 'yards' and 'rail'.

    Returns:
        pd.DataFrame: Updated `stockpiles_df` with an 'engines' column listing
            the assigned engine IDs.
    """
    stockpiles_df["engines"] = [[] for _ in range(stockpiles_df.shape[0])]

    for idx, stockpile in stockpiles_df.iterrows():
        assigned_engines = [
            eng_row["id"]
            for _, eng_row in engines_df.iterrows()
            if stockpile["yard"] in eng_row["yards"]
            and eng_row["rail"] in stockpile["rails"]
        ]
        stockpiles_df.at[idx, "engines"] = assigned_engines

    return stockpiles_df


def extract_quality_ini_values(
    stockpiles_df: pd.DataFrame, quality_prefix: str = "qualityIni"
) -> pd.DataFrame:
    """Extract initial quality values from nested dictionaries in the
    stockpiles DataFrame.

    This function processes a DataFrame containing stockpile information,
    specifically focusing on columns that start with a specified prefix
    related to quality parameters. It extracts the 'parameter' and 'value'
    from these nested dictionaries and adds them as new columns to the
    original DataFrame. The original quality columns are then removed from
    the DataFrame, resulting in a cleaner structure that separates quality
    data into individual columns.

    Args:
        stockpiles_df (pd.DataFrame): A `pandas.DataFrame` containing stockpile information,
            including quality parameters.
        quality_prefix (str?): Prefix used in column names for quality-related
            information. Defaults to 'qualityIni'.

    Returns:
        pd.DataFrame: Updated `stockpiles_df` with quality parameters extracted as individual
            columns.
    """
    quality_cols = stockpiles_df.columns[
        stockpiles_df.columns.str.startswith(quality_prefix)
    ]
    quality_ini_dict = {}

    for column in quality_cols:
        for idx, row in stockpiles_df.iterrows():
            parameter = row[column]["parameter"]
            value = row[column]["value"]
            quality_ini_dict.setdefault(parameter, []).append(value)

    # Add the extracted quality values as new columns and drop the original quality columns
    stockpiles_df = pd.concat(
        [stockpiles_df, pd.DataFrame(quality_ini_dict, index=stockpiles_df.index)],
        axis=1,
    ).drop(columns=quality_cols, errors="ignore")

    return stockpiles_df


def process_stockpiles_and_engines(
    stockpiles_df: pd.DataFrame, engines_df: pd.DataFrame
) -> pd.DataFrame:
    """Process stockpiles and engines by assigning engines to stockpiles and
    extracting initial quality values.

    This function takes two pandas DataFrames as input: one containing
    information about stockpiles and the other containing information about
    engines. It processes these DataFrames by assigning engines to the
    respective stockpiles based on certain criteria and then extracts the
    initial quality values from the stockpiles. The result is a modified
    DataFrame that includes the assigned engines and the extracted quality
    values.

    Args:
        stockpiles_df (pd.DataFrame): A `pandas.DataFrame` containing stockpile information.
        engines_df (pd.DataFrame): A `pandas.DataFrame` containing engine information.

    Returns:
        pd.DataFrame: Processed `stockpiles_df` with engines assigned and quality values
            extracted.
    """
    stockpiles_df = assign_engines_to_stockpiles(stockpiles_df, engines_df)
    stockpiles_df = extract_quality_ini_values(stockpiles_df)
    return stockpiles_df


def travel_time(grp: pd.DataFrame) -> pd.DataFrame:
    """Calculate the travel time between consecutive events within a group.

    This function computes the duration between the end of one event and the
    start of the next event in a grouped DataFrame. If the group consists of
    only one event, the travel time is set to 0. The function expects the
    DataFrame to be pre-grouped by a relevant key and to contain at least
    'start_time' and 'end_time' columns. The resulting DataFrame will have a
    single column named 'travel_time' that reflects the calculated travel
    times, with the index matching that of the input DataFrame.

    Args:
        grp (pd.DataFrame): A `pandas.DataFrame` containing at least 'start_time' and 'end_time'

    Returns:
        pd.DataFrame: A `pandas.DataFrame` with a single column 'travel_time',
        containing the calculated travel times between consecutive events.
        The index of the returned DataFrame matches the input DataFrame.
    """
    if len(grp) == 1:
        return pd.DataFrame(
            {"travel_time": [grp["start_time"].values[0]]}, index=grp.index
        )
    grp = grp.sort_values(["end_time"])
    end_time = None
    res = []
    for _, row in grp.iterrows():
        _end_time = row["end_time"]
        if end_time is None:
            res.append(row["start_time"])
        else:
            res.append(row["start_time"] - end_time)
        end_time = _end_time
    return pd.DataFrame({"travel_time": res}, index=grp.index)


def generate_excel_from_json_in_out(
    json_input_path: str, json_output_path: str, excel_path: str
):
    """Generate an Excel file from JSON input and output data.

    This function reads JSON data from specified input and output files,
    processes the data to create various DataFrames, and then compiles the
    results into an Excel file. The function handles multiple aspects of the
    data, including stockpiles, engines, and operations, and formats the
    final output for easy analysis. The resulting Excel file is saved at the
    specified path.

    Args:
        json_input_path (str): The file path to the input JSON file containing instance data.
        json_output_path (str): The file path to the output JSON file containing results data.
        excel_path (str): The file path where the generated Excel file will be saved.

    Returns:
        None: This function does not return a value; it generates an Excel file as
            output.
    """

    Path(excel_path).parent.mkdir(exist_ok=True, parents=True)
    excel_path = str(Path(excel_path).with_suffix(".xlsx"))

    # Load JSON files
    # Input file
    with open(json_input_path) as fh:
        instance_data = json.load(fh)

    # Output file
    with open(json_output_path) as fh:
        output_data = json.load(fh)

    # Convert `instance_1.json` to dataframe
    info_df = pd.DataFrame(
        [instance_data["info"]], columns=["Instance_Name", "Capacity", "Yard"]
    )
    engines_df = pd.DataFrame(instance_data["engines"])
    stockpiles_df = pd.DataFrame(instance_data["stockpiles"])
    stockpiles_quality_df = pd.json_normalize(
        stockpiles_df.pop("qualityIni"), sep="_"
    ).add_prefix("qualityIni_")
    stockpiles_df = pd.concat([stockpiles_df, stockpiles_quality_df], axis=1)
    stockpiles_df = process_stockpiles_and_engines(stockpiles_df, engines_df)

    inputs_df = pd.DataFrame(instance_data["inputs"])
    inputs_quality_df = pd.json_normalize(inputs_df.pop("quality"), sep="_").add_prefix(
        "quality_"
    )
    inputs_df = pd.concat([inputs_df, inputs_quality_df], axis=1)

    # Convert instance_1.json to DataFrames
    outputs_df = pd.DataFrame(instance_data["outputs"])
    outputs_quality_df = pd.json_normalize(
        outputs_df.pop("quality"), sep="_"
    ).add_prefix("quality_")
    outputs_df = pd.concat([outputs_df, outputs_quality_df], axis=1)

    # Explode the outputs_df
    outputs_df = explode_quality_rows(outputs_df, quality_col_prefix="quality_").drop(
        columns=["time"], errors="ignore"
    )
    distances_travel_df = pd.DataFrame(instance_data["distancesTravel"])
    time_travel_df = pd.DataFrame(instance_data["timeTravel"])

    time_travel_df.columns += 1
    time_travel_df.index += 1

    distances_travel_df.columns += 1
    distances_travel_df.index += 1

    from_to_list = []
    for col in time_travel_df.columns:
        for idx in time_travel_df.index:
            from_to_list.append([f"{col} -> {idx}", time_travel_df.loc[idx, col]])
    from_to_df = pd.DataFrame(from_to_list, columns=["from_to", "duration"])

    engines_df[from_to_df["from_to"].to_list()] = -1
    engines_df[from_to_df["from_to"].to_list()] = engines_df[
        from_to_df["from_to"].to_list()
    ].astype(float)

    for idx, row in engines_df.iterrows():
        yards = row["yards"]
        rail = row["rail"]
        stockpiles = []
        for _, stockpile_row in stockpiles_df.iterrows():
            if stockpile_row["yard"] in yards and rail in stockpile_row["rails"]:
                stockpiles.append(stockpile_row["id"])
        for start_stockpile in stockpiles:
            for end_stockpile in stockpiles:
                column_name = f"{start_stockpile} -> {end_stockpile}"
                duration = from_to_df.loc[
                    from_to_df["from_to"] == column_name, "duration"
                ].values[0]
                engines_df.loc[engines_df.index == idx, column_name] = duration

    engines_df[from_to_df["from_to"].to_list()] = engines_df[
        from_to_df["from_to"].to_list()
    ].replace(-1, "")

    # Convert out_1.json to DataFrames
    objective_df = pd.DataFrame(
        [{"Objective": output_data["objective"], "Gap": output_data["gap"][0]}]
    )
    stacks_df = pd.DataFrame(output_data["stacks"])
    reclaims_df = pd.DataFrame(output_data["reclaims"])

    outputs_df_out = pd.DataFrame(output_data["outputs"])
    outputs_quality_df_out = pd.json_normalize(
        outputs_df_out.pop("quality"), sep="_"
    ).add_prefix("quality_")
    outputs_df_out = pd.concat([outputs_df_out, outputs_quality_df_out], axis=1)
    outputs_df_out = explode_quality_rows(outputs_df_out, quality_col_prefix="quality_")

    if not stacks_df.empty:
        stacks_df["end_time"] = stacks_df["start_time"] + stacks_df["duration"]
        stacks_df["operation"] = "stack"

    if not reclaims_df.empty:
        reclaims_df["end_time"] = reclaims_df["start_time"] + reclaims_df["duration"]
        reclaims_df["operation"] = "reclaim"

    operations_df = (
        pd.concat([xdf for xdf in [stacks_df, reclaims_df] if not xdf.empty])
        .astype({"weight": int})
        .sort_values(["engine", "start_time"])
        .assign(
            travel_time=lambda xdf: (
                xdf.groupby("engine", as_index=False).apply(travel_time)
            )["travel_time"].reset_index(level=0, drop=True)
        )
    )

    stockpiles_final_df = (
        stockpiles_df.merge(
            operations_df.rename(columns={"weight": "weightFinal"})
            .groupby("stockpile")["weightFinal"]
            .sum(),
            left_on="id",
            right_index=True,
            how="left",
        )
        .fillna({"weightFinal": 0})
        .astype({"weightFinal": int})
        .assign(weightFinal=lambda xdf: xdf["weightIni"] - xdf["weightFinal"])
    )

    quality_cols = stockpiles_df.columns.intersection(
        ["Fe", "SiO2", "Al2O3", "P", "+31.5", "-3"]
    ).to_list()
    operations_df = operations_df.merge(
        stockpiles_df.rename(columns={"id": "stockpile"})[
            ["stockpile", "weightIni", *quality_cols]
        ],
        on="stockpile",
        how="left",
    ).assign(weightFinal=lambda xdf: xdf["weightIni"] - xdf["weight"])

    final_output_row = [
        operations_df["weight"].sum(),
        "output",
        operations_df["engine"].unique().tolist(),
        operations_df["start_time"].min(),
        operations_df["end_time"].max(),
        1,
        operations_df["end_time"].max(),
        "output_stack",
        operations_df["travel_time"].sum(),
        operations_df["weightFinal"].sum(),
    ]
    for quality_col in quality_cols:
        final_output_row.append(
            (
                (
                    operations_df["weight"]
                    * operations_df[quality_col]
                    / operations_df["weight"].sum()
                ).sum()
            )
        )
    operations_df = pd.concat(
        [
            operations_df,
            pd.DataFrame(
                {
                    col: [value]
                    for col, value in zip(operations_df.columns, final_output_row)
                }
            ),
        ],
        axis=0,
    )

    create_formatted_excel(
        stockpiles_df,
        engines_df,
        operations_df.reset_index(drop=True),
        outputs_df_out,
        file_name=excel_path,
    )


if __name__ == "__main__":
    generate_excel_from_json_in_out(
        "../tests/instance_8.json",
        "../out/json/out_8.json",
        "../out/excel/out_8.xlsx",
    )
